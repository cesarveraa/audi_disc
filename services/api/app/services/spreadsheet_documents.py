from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


AUDI_RED = "E4002B"
INK = "111827"
WHITE = "FFFFFF"


def _money(centavos: int) -> float:
    return round(centavos / 100, 2)


def _finish_workbook(workbook: Workbook) -> bytes:
    for worksheet in workbook.worksheets:
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color=WHITE)
            cell.fill = PatternFill("solid", fgColor=INK)
        for column_cells in worksheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 2, 12), 42)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def products_inventory_xlsx(products: list[dict]) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Productos"
    worksheet.append(
        [
            "ID",
            "Nombre",
            "Marca",
            "SKU",
            "Categoria",
            "Stock",
            "Stock minimo",
            "Precio venta Bs",
            "Precio compra Bs",
            "Utilidad Bs",
            "Margen %",
            "Estado",
        ]
    )
    for product in products:
        worksheet.append(
            [
                product.get("id", ""),
                product.get("nombre", ""),
                product.get("marca") or "",
                product.get("sku") or "",
                product.get("categoria") or "",
                int(product.get("cantidad", 0)),
                int(product.get("stockMinimo", 0)),
                _money(int(product.get("precioVentaCentavos", 0))),
                _money(int(product.get("precioCompraCentavos", 0))),
                _money(int(product.get("utilidadCentavos", 0))),
                float(product.get("margenPorcentaje", 0)),
                "Activo" if product.get("estado", True) else "Inactivo",
            ]
        )
    return _finish_workbook(workbook)


def sales_history_xlsx(history: dict) -> bytes:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Resumen"
    summary.append(["Metrica", "Valor"])
    summary.append(["Desde", history.get("dateFrom", "")])
    summary.append(["Hasta", history.get("dateTo", "")])
    summary.append(["Total ventas", int(history.get("cantidadVentas", 0))])
    summary.append(["Total vendido Bs", _money(int(history.get("totalCentavos", 0)))])
    summary.append(["Utilidad Bs", _money(int(history.get("utilidadCentavos", 0)))])
    summary.append(["Margen %", float(history.get("margenPorcentaje", 0))])

    sales = workbook.create_sheet("Ventas")
    sales.append(["ID", "Fecha", "Hora", "Metodo", "Total Bs", "Recibido Bs", "Cambio Bs", "Estado"])
    for sale in history.get("ventas", []):
        sales.append(
            [
                sale.get("id", ""),
                sale.get("fechaLocal", ""),
                sale.get("horaLocal", ""),
                sale.get("metodo", ""),
                _money(int(sale.get("totalCentavos", 0))),
                _money(int(sale.get("recibidoCentavos", 0))),
                _money(int(sale.get("cambioCentavos", 0))),
                "Activa" if sale.get("estado", True) else "Anulada",
            ]
        )

    items = workbook.create_sheet("Detalle")
    items.append(["Venta", "Producto", "Marca", "SKU", "Categoria", "Cantidad", "Precio Bs", "Subtotal Bs", "Utilidad Bs"])
    for sale in history.get("ventas", []):
        for item in sale.get("productos", []):
            items.append(
                [
                    sale.get("id", ""),
                    item.get("nombre", ""),
                    item.get("marca") or "",
                    item.get("sku") or "",
                    item.get("categoria") or "",
                    int(item.get("cantidad", 0)),
                    _money(int(item.get("precioVendidoCentavos") or item.get("precioVentaCentavos") or 0)),
                    _money(int(item.get("subtotalCentavos", 0))),
                    _money(int(item.get("utilidadCentavos", 0))),
                ]
            )

    return _finish_workbook(workbook)
